#!/usr/bin/env python3
"""Generate valid Excel templates with Named Ranges from YAML specification."""

import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname, absolute_coordinate
import yaml
from pathlib import Path
import hashlib

def create_templates():
    """Create Cover.xlsx and Estimate.xlsx with proper Named Ranges."""

    # Load Named Ranges from YAML
    yaml_path = Path("KIS/Templates/NamedRanges.yaml")
    with open(yaml_path, 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)

    ranges = config.get('ranges', [])

    # Create Cover.xlsx
    print("Creating Cover.xlsx...")
    wb_cover = openpyxl.Workbook()
    ws = wb_cover.active
    ws.title = "Cover"

    # Sample values
    ws['B3'] = "KIS Electrical Installation"
    ws['B4'] = "Sample Client Co."
    ws['B5'] = "2024-01-01"
    ws['B6'] = "PRJ-2024-001"
    ws['B7'] = "Project Manager"
    ws['B60'] = "Prepared By"
    ws['D60'] = "2024-01-01"

    # Add Named Ranges for Cover
    for r in ranges:
        if r['sheet'] == 'Cover':
            ref = f"{quote_sheetname(r['sheet'])}!{absolute_coordinate(r['ref'])}"
            wb_cover.defined_names[r['name']] = ref

    cover_path = Path("KIS/Templates/Cover.xlsx")
    wb_cover.save(cover_path)
    wb_cover.close()

    # Create Estimate.xlsx
    print("Creating Estimate.xlsx...")
    wb_estimate = openpyxl.Workbook()
    ws = wb_estimate.active
    ws.title = "Estimate"

    # Sample values
    ws['A10'] = "ITEMS"
    ws['B10'] = "Description"
    ws['C10'] = "Qty"
    ws['D10'] = "Unit"
    ws['E10'] = "Unit Price"
    ws['F10'] = "Discount"
    ws['G10'] = "Tax"
    ws['H10'] = "Total"

    # Sample item row
    ws['A11'] = "1"
    ws['B11'] = "Main Panel"
    ws['C11'] = 1
    ws['D11'] = "EA"
    ws['E11'] = 500000
    ws['H11'] = 500000

    # Totals (numeric format)
    ws['H52'] = 0  # Net
    ws['H53'] = 0  # Discount
    ws['H54'] = 0  # Subtotal
    ws['H55'] = 0  # VAT
    ws['H56'] = 0  # Total

    # Add Named Ranges for Estimate
    for r in ranges:
        if r['sheet'] == 'Estimate':
            if ':' in r['ref']:  # Range
                ref = f"{quote_sheetname(r['sheet'])}!{r['ref']}"
            else:  # Single cell
                ref = f"{quote_sheetname(r['sheet'])}!{absolute_coordinate(r['ref'])}"
            wb_estimate.defined_names[r['name']] = ref

    estimate_path = Path("KIS/Templates/Estimate.xlsx")
    wb_estimate.save(estimate_path)
    wb_estimate.close()

    # Calculate hashes and sizes
    results = []
    for path in [cover_path, estimate_path]:
        size = path.stat().st_size
        with open(path, 'rb') as f:
            hash_val = hashlib.sha256(f.read()).hexdigest()
        results.append({
            'file': path.name,
            'size': size,
            'hash': hash_val[:16],
            'sheets': 1,
            'named_ranges': len([r for r in ranges if r['sheet'] in path.stem])
        })

    return results, len(ranges)

if __name__ == "__main__":
    try:
        results, total_ranges = create_templates()
        print("\nTemplate Generation Complete:")
        for r in results:
            print(f"  {r['file']}: {r['size']} bytes, {r['named_ranges']} ranges")
        print(f"Total Named Ranges: {total_ranges}")
    except Exception as e:
        print(f"Error: {e}")
        exit(1)