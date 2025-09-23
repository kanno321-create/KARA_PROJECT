# system_estimate_ui.py
import os, json, time, re, traceback, threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

# ===== 기존 엔진 유지 사용 =====
from estimate_engine import compute_enclosure_size, build_estimate_lines

# ===== 선택 의존성 =====
_PIL_OK=_TESS_OK=_PDFPLUMBER_OK=_DOCX_OK=_PDF2IMG_OK=_DND_OK=True
try:
    from PIL import Image
except: _PIL_OK=False
try:
    import pytesseract
except: _TESS_OK=False
try:
    import pdfplumber
except: _PDFPLUMBER_OK=False
try:
    import docx
except: _DOCX_OK=False
try:
    from pdf2image import convert_from_path
except: _PDF2IMG_OK=False
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
except: _DND_OK=False

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from assistant_agent import AssistantAgent  # 에이전트

# === 표준 메일 라이브러리
import imaplib, email
from email.header import decode_header
from email.utils import parsedate_to_datetime

APP_TITLE = "한국산업 견적 도우미"
BASE_DIR = os.path.dirname(__file__)
DATA_DIR = os.path.join(BASE_DIR, "data")
EST_DIR  = os.path.join(DATA_DIR, "estimates")
CFG_PATH = os.path.join(DATA_DIR, "settings.json")
CLIENTS_JSON = os.path.join(DATA_DIR, "clients.json")
PRICES_DIR = os.path.join(DATA_DIR, "prices")
for d in (DATA_DIR, EST_DIR, PRICES_DIR):
    os.makedirs(d, exist_ok=True)

# === IMAP 서버 고정값 (네이버웍스/네이버 공용)
IMAP_HOST = "imap.naver.com"
IMAP_PORT = 993  # SSL

def save_json(path, data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_json(path, default=None):
    if not os.path.exists(path): return default
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def now_tag(): return time.strftime("%y%m%d_%H%M%S")
def slug(s):
    v = re.sub(r'[^가-힣a-zA-Z0-9_\-]+', '_', s or "")
    return v.strip('_') or "NONAME"
def safe_int(x, d=0):
    try: return int(str(x).strip())
    except: return d

# -------------------------
# OCR/문서 파서
# -------------------------
AMP_CHOICES = [15,20,30,40,50,60,75,100,125,150,175,200,225,250,300,350,400,500,600,630,700,800]
def _norm(t): return re.sub(r"[ \t]+"," ",t or "").replace("\r","\n")
def _extract_email(t):
    m=re.search(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", t or "")
    return m.group(0) if m else ""
def _extract_phone(t):
    m=re.search(r"(01[0-9]-?\d{3,4}-?\d{4}|0\d{1,2}-?\d{3,4}-?\d{4})", t or "")
    return m.group(0) if m else ""
def _extract_company(t):
    for key in ["업체명","회사","상호","법인명","거래처"]:
        m=re.search(key+r"\s*[:：]\s*([^\n]+)", t or "")
        if m:
            name=re.sub(r"[^가-힣A-Za-z0-9()_\- ]+","",m.group(1)).strip()
            return name[:60]
    mail=_extract_email(t or "")
    if mail:
        guess=re.sub(r"[\._\-0-9]+$","", mail.split("@")[0])
        if guess: return guess[:60]
    return ""

def _ocr_image(path, tesseract_path=""):
    if not (_PIL_OK and _TESS_OK): return ""
    try:
        if tesseract_path and os.path.exists(tesseract_path):
            pytesseract.pytesseract.tesseract_cmd=tesseract_path
        img=Image.open(path)
        return pytesseract.image_to_string(img, lang="kor+eng")
    except: return ""

def _pdf_to_text(path, tesseract_path=""):
    texts=[]
    if _PDFPLUMBER_OK:
        try:
            with pdfplumber.open(path) as pdf:
                for p in pdf.pages:
                    t=p.extract_text() or ""
                    if t.strip(): texts.append(t)
        except: pass
    joined="\n".join(texts)
    if joined.strip(): return joined
    if _PDF2IMG_OK and _PIL_OK and _TESS_OK:
        try:
            if tesseract_path and os.path.exists(tesseract_path):
                pytesseract.pytesseract.tesseract_cmd=tesseract_path
            imgs=convert_from_path(path, dpi=200)
            buf=[]
            for im in imgs[:10]:
                buf.append(pytesseract.image_to_string(im, lang="kor+eng"))
            return "\n".join(buf)
        except: return ""
    return ""

def _scan_specs(text):
    text = text or ""
    lines=[s.strip() for s in text.splitlines() if s.strip()]
    specs=[]
    for ln in lines:
        kind="MCCB" if re.search(r"\bMCCB\b", ln, re.I) else ("ELB" if (re.search(r"\bELB\b", ln, re.I) or "누전" in ln) else None)
        poles=None; amp=None; qty=1
        m=re.search(r"\b([234]P)\b", ln, re.I);     poles=m.group(1).upper() if m else None
        m2=re.search(r"(\d{2,3})\s*A\b", ln, re.I)
        if m2:
            av=safe_int(m2.group(1)); amp=f"{min(AMP_CHOICES,key=lambda x:abs(x-av))}A"
        mqty=re.search(r"(?:x|X|수량\s*)(\d+)\b|(\d+)\s*EA\b", ln, re.I)
        if mqty: qty=safe_int([g for g in mqty.groups() if g][0],1)
        if kind and poles and amp:
            specs.append({"종류":kind,"극수":poles,"용량":amp,"수량":str(qty)})
    # 2-line 조합
    for i in range(len(lines)-1):
        ln1,ln2=lines[i],lines[i+1]
        kind="MCCB" if (re.search(r"\bMCCB\b",ln1+ln2,re.I)) else ("ELB" if (re.search(r"\bELB\b",ln1+ln2,re.I) or "누전" in (ln1+ln2)) else None)
        poles=None; amp=None; qty=1
        m=re.search(r"\b([234]P)\b", ln1+" "+ln2, re.I); poles=m.group(1).upper() if m else None
        m2=re.search(r"(\d{2,3})\s*A\b", ln1+" "+ln2, re.I)
        if m2:
            av=safe_int(m2.group(1)); amp=f"{min(AMP_CHOICES,key=lambda x:abs(x-av))}A"
        mqty=re.search(r"(?:x|X|수량\s*)(\d+)\b|(\d+)\s*EA\b", ln1+" "+ln2, re.I)
        if mqty: qty=safe_int([g for g in mqty.groups() if g][0],1)
        if kind and poles and amp:
            specs.append({"종류":kind,"극수":poles,"용량":amp,"수량":str(qty)})
    return specs

def _split_main_branch(specs, text):
    text = text or ""
    for s in specs:
        patt=rf"{s['극수']}.*{s['용량']}"
        if re.search(r"메인|MAIN|주차단기", text, re.I) and re.search(patt, text, re.I):
            return s, [x for x in specs if x is not s]
    def score(sp):
        amp=safe_int(sp['용량'].replace("A",""),0)
        pole_rank={"4P":3,"3P":2,"2P":1}.get(sp['극수'],0)
        return (amp, pole_rank)
    if specs:
        main=sorted(specs, key=score, reverse=True)[0]
        b=specs.copy(); b.remove(main)
        return main, b
    return None, []

def extract_info_from_files(paths, tesseract_path=""):
    texts=[]
    for p in paths:
        ext=os.path.splitext(p)[1].lower()
        t=""
        if ext in [".png",".jpg",".jpeg",".bmp",".tif",".tiff"]:
            t=_ocr_image(p, tesseract_path)
        elif ext==".pdf":
            t=_pdf_to_text(p, tesseract_path)
        elif ext==".docx" and _DOCX_OK:
            try:
                doc=docx.Document(p)
                t="\n".join([pa.text for pa in doc.paragraphs])
            except: t=""
        else:
            try:
                with open(p,"r",encoding="utf-8") as f: t=f.read()
            except:
                try:
                    with open(p,"r",encoding="cp949") as f: t=f.read()
                except: t=os.path.basename(p)
        if t: texts.append(t)
    full=_norm("\n".join(texts))
    client={"업체명":_extract_company(full),"연락처":_extract_phone(full),"이메일":_extract_email(full)}
    specs=_scan_specs(full)
    main, branches=_split_main_branch(specs, full)
    return client, main, branches, full

# -------------------------
# Tk Base(드래그앤드롭 지원 여부)
# -------------------------
BaseTk = TkinterDnD.Tk if _DND_OK else tk.Tk

def _dh(s):
    parts = decode_header(s or "")
    out=[]
    for txt, enc in parts:
        if isinstance(txt, bytes):
            out.append(txt.decode(enc or "utf-8", errors="ignore"))
        else:
            out.append(txt)
    return "".join(out).strip()

def _msg_summary(msg):
    subj=_dh(msg.get("Subject",""))
    frm=_dh(msg.get("From",""))
    try:
        dt=parsedate_to_datetime(msg.get("Date"))
        date_s = dt.strftime("%Y-%m-%d %H:%M")
    except:
        date_s = msg.get("Date","")
    atts=0
    for part in msg.walk():
        if (part.get_content_disposition() or "").lower() == "attachment":
            atts+=1
    return date_s, frm, subj, atts

class App(BaseTk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1350x900")
        self.cfg=load_json(CFG_PATH, default={
            "imap": {"host":"imap.naver.com","port":993,"user":"","pass":"","ssl":True,"ok":False},
            "smtp":{"host":"","port":587,"user":"","pass":"","sender":""},
            "naver_works":{"id":"","api_key":""},
            "chatgpt":{"url":"https://api.openai.com/v1/chat/completions","api_key":"","ok":False,"model":"gpt-4o"},
            "claude":{"url":"https://api.anthropic.com/v1/messages","api_key":"","ok":False,"model":"claude-3-7-sonnet-20250219"},
            "tesseract_path":""
        })

        self._init_state()
        self.agent = AssistantAgent(base_dir=BASE_DIR, data_dir=DATA_DIR)

        self.style = ttk.Style()
        try: self.style.theme_use("clam")
        except: pass

        self.nb = ttk.Notebook(self); self.nb.pack(fill="both", expand=True)
        self._build_tab_estimate()
        self._build_tab_ai()
        self._build_tab_email()
        self._build_tab_repo()
        self._build_tab_prices()
        self._build_tab_clients()
        self._build_tab_shipping()
        self._build_tab_reports()
        self._build_tab_settings()
        self.nb.select(0)

    def _init_state(self):
        self.branches=[]; self.accessories=[]
        self.lines=[]; self.last_enclosure=None; self.last_main=None; self.last_client=None
        # 메일 캐시
        self.mail_cache = {}

    # ========= [탭1] 견적 =========
    def _build_tab_estimate(self):
        tab = ttk.Frame(self.nb, padding=8); self.nb.add(tab, text="견적")
        left = ttk.Frame(tab); left.pack(side="left", fill="y")

        # 고객
        g0=ttk.LabelFrame(left, text="고객 정보", padding=8); g0.pack(fill="x", pady=(0,8))
        ttk.Label(g0, text="업체명").grid(row=0,column=0,sticky="w")
        ttk.Label(g0, text="연락처").grid(row=1,column=0,sticky="w")
        ttk.Label(g0, text="이메일").grid(row=2,column=0,sticky="w")
        self.ent_client_name=ttk.Entry(g0,width=26); self.ent_client_name.grid(row=0,column=1,padx=4)
        self.ent_client_phone=ttk.Entry(g0,width=26); self.ent_client_phone.grid(row=1,column=1,padx=4)
        self.ent_client_email=ttk.Entry(g0,width=26); self.ent_client_email.grid(row=2,column=1,padx=4)
        ttk.Button(g0,text="이메일 확인 후 불러오기", command=self._fill_client_from_selected_email).grid(row=0,column=2,rowspan=3,padx=(8,0))

        # 외함 (개선)
        g1=ttk.LabelFrame(left, text="외함", padding=8); g1.pack(fill="x", pady=(0,8))
        ttk.Label(g1,text="설치").grid(row=0,column=0,sticky="e")
        self.cbo_enc_place=ttk.Combobox(g1, values=["옥내","옥외","옥내자립","옥외자립","전주부착형"],
                                        width=10, state="readonly"); self.cbo_enc_place.set("옥내")
        self.cbo_enc_place.grid(row=0,column=1,padx=4,sticky="w")

        ttk.Label(g1,text="함종류").grid(row=1,column=0,sticky="e")
        self.cbo_enc_kind=ttk.Combobox(g1, values=["기성함","주문제작함","계량기함","CT계량기함","FRP 박스","하이박스"],
                                       width=12, state="readonly"); self.cbo_enc_kind.set("기성함")
        self.cbo_enc_kind.grid(row=1,column=1,padx=4,sticky="w")

        ttk.Label(g1,text="재질").grid(row=2,column=0,sticky="e")
        self.cbo_enc_mat=ttk.Combobox(g1, values=["STEEL 1.0T","STEEL 1.6T","STEEL 2.0T","SUS201 1.0T","SUS201 1.2T","SUS201 1.5T","SUS304 1.2T","SUS304 1.5T","SUS304 2.0T"],
                                      width=16, state="readonly"); self.cbo_enc_mat.set("STEEL 1.6T")
        self.cbo_enc_mat.grid(row=2,column=1,padx=4,sticky="w")

        ttk.Label(g1,text="기타요청").grid(row=3,column=0,sticky="e")
        self.ent_enc_misc=ttk.Entry(g1, width=26)
        self.ent_enc_misc.grid(row=3,column=1,padx=4,sticky="w")

        ttk.Label(g1,text="주문제작 단가").grid(row=4,column=0,sticky="e")
        self.ent_enc_custom_price=ttk.Entry(g1, width=12)
        self.ent_enc_custom_price.grid(row=4,column=1,padx=4,sticky="w")
        ttk.Label(g1,text="(함종류가 '주문제작함'일 때만 적용)").grid(row=4,column=2,sticky="w")

        self.lbl_base_auto = ttk.Label(g1, text="※ 자립형(옥내/옥외자립)은 베이스 포함으로 자동 처리")
        self.lbl_base_auto.grid(row=5, column=0, columnspan=3, sticky="w", pady=(4,0))

        # 메인
        POLES=["2P","3P","4P"]
        AMPS=["15A","20A","30A","40A","50A","60A","75A","100A","125A","150A","175A","200A","225A","250A","300A","350A","400A","500A","600A","630A","700A","800A"]
        g2=ttk.LabelFrame(left, text="메인 차단기", padding=8); g2.pack(fill="x", pady=(0,8))
        ttk.Label(g2,text="종류").grid(row=0,column=0,sticky="w")
        ttk.Label(g2,text="극수").grid(row=1,column=0,sticky="w")
        ttk.Label(g2,text="용량").grid(row=2,column=0,sticky="w")
        ttk.Label(g2,text="수량").grid(row=3,column=0,sticky="w")
        self.cbo_main_kind =ttk.Combobox(g2, values=["MCCB","ELB"], state="readonly", width=8); self.cbo_main_kind.set("MCCB"); self.cbo_main_kind.grid(row=0,column=1,padx=4)
        self.cbo_main_poles=ttk.Combobox(g2, values=POLES, state="readonly", width=8); self.cbo_main_poles.set("4P"); self.cbo_main_poles.grid(row=1,column=1,padx=4)
        self.cbo_main_amp  =ttk.Combobox(g2, values=AMPS, state="readonly", width=8); self.cbo_main_amp.set("100A"); self.cbo_main_amp.grid(row=2,column=1,padx=4)
        self.ent_main_qty  =ttk.Entry(g2, width=10); self.ent_main_qty.insert(0,"1"); self.ent_main_qty.grid(row=3,column=1,padx=4)

        ttk.Label(g2,text="브랜드").grid(row=0,column=2,sticky="w")
        self.cbo_main_brand = ttk.Combobox(g2, values=["상도","LS","대륙","비츠로","현대"], state="readonly", width=8)
        self.cbo_main_brand.set("상도")
        self.cbo_main_brand.grid(row=0,column=3,padx=4)

        self.var_spd=tk.BooleanVar(value=False); ttk.Checkbutton(g2, text="SPD 포함", variable=self.var_spd).grid(row=4,column=0,columnspan=2,sticky="w")

        # 분기
        g3=ttk.LabelFrame(left, text="분기 차단기", padding=8); g3.pack(fill="x", pady=(0,8))
        self.cbo_b_kind =ttk.Combobox(g3, values=["MCCB","ELB"], state="readonly", width=8); self.cbo_b_kind.set("ELB"); self.cbo_b_kind.grid(row=0,column=0,padx=2)
        self.cbo_b_poles=ttk.Combobox(g3, values=POLES, state="readonly", width=8); self.cbo_b_poles.set("2P"); self.cbo_b_poles.grid(row=0,column=1,padx=2)
        self.cbo_b_amp  =ttk.Combobox(g3, values=AMPS, state="readonly", width=8); self.cbo_b_amp.set("30A"); self.cbo_b_amp.grid(row=0,column=2,padx=2)
        self.ent_b_qty  =ttk.Entry(g3, width=6); self.ent_b_qty.insert(0,"4"); self.ent_b_qty.grid(row=0,column=3,padx=2)
        ttk.Button(g3, text="분기 추가", command=self._add_branch).grid(row=0,column=4,padx=4)
        ttk.Button(g3, text="선택 삭제", command=self._remove_branch).grid(row=0,column=5,padx=2)
        self.lst_branches=tk.Listbox(g3, height=8, width=45); self.lst_branches.grid(row=1,column=0,columnspan=6,sticky="we",pady=(6,0))

        # 부속자재
        gAcc=ttk.LabelFrame(left, text="부속자재", padding=8); gAcc.pack(fill="x", pady=(0,8))
        MAGNETS=["MC-22","MC-32","MC-40","MC-50","MC-65","MC-75","MC-80"]
        ttk.Label(gAcc,text="마그네트").grid(row=0,column=0,sticky="w")
        self.cbo_mag=ttk.Combobox(gAcc, values=MAGNETS, width=12, state="readonly"); self.cbo_mag.grid(row=0,column=1,padx=2)
        self.ent_mag_qty=ttk.Entry(gAcc, width=6); self.ent_mag_qty.insert(0,"1"); self.ent_mag_qty.grid(row=0,column=2,padx=2)
        ttk.Button(gAcc,text="추가",command=self._add_acc_mag).grid(row=0,column=3,padx=4)

        CAP_VOLT=["220V","380V"]; CAP_SIZE=["10","15","20","30","40","50","60","75","100","150","175","200","250","300","400","500","1000"]
        ttk.Label(gAcc,text="콘덴서").grid(row=1,column=0,sticky="w")
        self.cbo_cap_v=ttk.Combobox(gAcc, values=CAP_VOLT, width=6, state="readonly"); self.cbo_cap_v.grid(row=1,column=1,padx=2,sticky="w")
        self.cbo_cap_s=ttk.Combobox(gAcc, values=CAP_SIZE, width=6, state="readonly"); self.cbo_cap_s.grid(row=1,column=2,padx=2,sticky="w")
        self.ent_cap_qty=ttk.Entry(gAcc, width=6); self.ent_cap_qty.insert(0,"1"); self.ent_cap_qty.grid(row=1,column=3,padx=2)
        ttk.Button(gAcc,text="추가",command=self._add_acc_cap).grid(row=1,column=4,padx=4)

        METERS=["V/A-meter","3상계량기(피에스텍)","단상계량기(피에스텍)","CT계량기(피에스텍)","3상계량기(LS)","단상계량기(LS)","CT계량기(LS)"]
        ttk.Label(gAcc,text="계측기").grid(row=2,column=0,sticky="w")
        self.cbo_meter=ttk.Combobox(gAcc, values=METERS, width=24, state="readonly"); self.cbo_meter.grid(row=2,column=1,columnspan=2,padx=2,sticky="w")
        self.ent_meter_qty=ttk.Entry(gAcc, width=6); self.ent_meter_qty.insert(0,"1"); self.ent_meter_qty.grid(row=2,column=3,padx=2)
        ttk.Button(gAcc,text="추가",command=self._add_acc_meter).grid(row=2,column=4,padx=4)

        ttk.Label(gAcc,text="기타부자재").grid(row=3,column=0,sticky="w")
        self.ent_misc_name=ttk.Entry(gAcc, width=24); self.ent_misc_name.grid(row=3,column=1,columnspan=2,padx=2,sticky="w")
        self.ent_misc_qty=ttk.Entry(gAcc, width=6); self.ent_misc_qty.insert(0,"1"); self.ent_misc_qty.grid(row=3,column=3,padx=2)
        ttk.Button(gAcc,text="추가",command=self._add_acc_misc).grid(row=3,column=4,padx=4)

        self.lst_acc=tk.Listbox(gAcc, height=7, width=52); self.lst_acc.grid(row=4,column=0,columnspan=5,sticky="we",pady=(6,2))
        ttk.Button(gAcc, text="선택 삭제", command=self._del_acc).grid(row=5,column=0,pady=(2,0))
        ttk.Button(gAcc, text="전체 초기화", command=self._clear_acc).grid(row=5,column=1,pady=(2,0))

        # 가운데 버튼
        mid=ttk.Frame(tab,padding=8); mid.pack(side="left", fill="y")
        ttk.Button(mid,text="시스템 견적 생성", command=self._make_system_estimate).pack(fill="x",pady=(0,6))
        ttk.Button(mid,text="AI 견적 생성", command=self._make_ai_estimate).pack(fill="x",pady=(0,6))
        ttk.Button(mid,text="화면 지우기", command=self._clear_output).pack(fill="x")

        # 우측 출력(표 + 합계)
        right=ttk.Frame(tab,padding=8); right.pack(side="left", fill="both", expand=True)

        sumbar = ttk.Frame(right); sumbar.pack(fill="x")
        self.var_est_title = tk.StringVar(value="(견적 없음)")
        ttk.Label(sumbar, textvariable=self.var_est_title).pack(side="left")
        self.var_est_total = tk.StringVar(value="합계: 0 원")
        ttk.Label(sumbar, textvariable=self.var_est_total).pack(side="right")

        cols=("no","name","spec","unit","qty","price","amount")
        self.grid_est = ttk.Treeview(right, columns=cols, show="headings", height=24)
        headers = ["No","품명","규격","단위","수량","단가","금액"]
        widths  = [50,260,300,60,70,120,140]
        for c,h,w in zip(cols,headers,widths):
            self.grid_est.heading(c, text=h)
            self.grid_est.column(c, width=w, anchor=("e" if c in ("qty","price","amount","no") else "w"))
        self.grid_est.pack(fill="both", expand=True, pady=(6,0))

        btns = ttk.Frame(right); btns.pack(fill="x", pady=(6,0))
        ttk.Button(btns, text="엑셀 저장(회사 양식)", command=self._save_to_company_excel).pack(side="left")
        ttk.Button(btns, text="CSV 저장", command=self._repo_save_csv).pack(side="left", padx=6)

    def _append(self,msg):
        # (이제 텍스트 영역 사용 안 함) – 남겨두되 호출 안함
        pass
    def _clear_output(self):
        # 표 초기화
        for iid in self.grid_est.get_children():
            self.grid_est.delete(iid)
        self.var_est_title.set("(견적 없음)")
        self.var_est_total.set("합계: 0 원")

    def _mock_fetch_email_contact(self):
        # (사용 안 함: 실제버전으로 대체)
        pass

    # 분기/부자재
    def _add_branch(self):
        k=self.cbo_b_kind.get().strip(); p=self.cbo_b_poles.get().strip(); a=self.cbo_b_amp.get().strip(); q=self.ent_b_qty.get().strip() or "1"
        try:
            if int(q)<=0: raise ValueError
        except:
            messagebox.showwarning("수량 오류","1 이상 숫자"); return
        self.branches.append({"종류":k,"극수":p,"용량":a,"수량":q})
        self.lst_branches.insert(tk.END, f"{k} | {p} {a} | 수량 {q}")
    def _remove_branch(self):
        sel=self.lst_branches.curselection()
        if not sel: return
        idx=sel[0]; self.lst_branches.delete(idx); del self.branches[idx]
    def _push_acc(self,name,qty):
        try:
            q=int(str(qty).strip() or "1")
            if q<=0: raise ValueError
        except:
            messagebox.showwarning("수량 오류","1 이상 숫자"); return
        self.accessories.append({"name":name,"qty":q})
        self.lst_acc.insert(tk.END, f"{name} x {q}")
    def _add_acc_mag(self):
        n=self.cbo_mag.get().strip()
        if n: self._push_acc(f"마그네트 {n}", self.ent_mag_qty.get())
    def _add_acc_cap(self):
        v=self.cbo_cap_v.get().strip(); s=self.cbo_cap_s.get().strip()
        if v and s: self._push_acc(f"콘덴서 {v} {s}μF", self.ent_cap_qty.get())
    def _add_acc_meter(self):
        m=self.cbo_meter.get().strip()
        if m: self._push_acc(m, self.ent_meter_qty.get())
    def _add_acc_misc(self):
        t=self.ent_misc_name.get().strip()
        if not t: messagebox.showwarning("입력 필요","이름 입력"); return
        self._push_acc(t, self.ent_misc_qty.get()); self.ent_misc_name.delete(0,tk.END)
    def _del_acc(self):
        sel=self.lst_acc.curselection()
        if not sel: return
        idx=sel[0]; self.lst_acc.delete(idx); del self.accessories[idx]
    def _clear_acc(self):
        self.accessories.clear(); self.lst_acc.delete(0,tk.END)

    def _render_estimate_table(self, lines: list):
        for iid in self.grid_est.get_children():
            self.grid_est.delete(iid)
        total = 0
        for L in lines:
            try:
                amt = int(float(L.get("금액",0)))
            except:
                amt = 0
            total += amt
            self.grid_est.insert("", "end", values=(
                L.get("no",""),
                L.get("품명",""),
                L.get("규격",""),
                L.get("단위",""),
                L.get("수량",""),
                L.get("단가",""),
                L.get("금액",""),
            ))
        self.var_est_total.set(f"합계: {total:,} 원")

    def _build_and_show_estimate(self, mode="system"):
        place = self.cbo_enc_place.get()
        kind  = self.cbo_enc_kind.get()
        mat   = self.cbo_enc_mat.get()
        misc  = self.ent_enc_misc.get().strip()
        custom_price = self.ent_enc_custom_price.get().strip()
        has_base = place in ("옥내자립","옥외자립")
        enclosure={
            "설치": place,
            "함종류": kind,
            "외함 재질": mat,
            "베이스 유무": ("있음" if has_base else "없음"),
            "기타요청": misc,
        }
        if kind=="주문제작함" and custom_price:
            enclosure["주문제작_단가"] = custom_price

        main={"종류": self.cbo_main_kind.get(),
              "극수": self.cbo_main_poles.get(),
              "용량": self.cbo_main_amp.get(),
              "브랜드": self.cbo_main_brand.get(),
              "수량": self.ent_main_qty.get().strip() or "1"}
        client={"업체명": self.ent_client_name.get().strip(), "연락처": self.ent_client_phone.get().strip(), "이메일": self.ent_client_email.get().strip()}
        try:
            W,H,D = compute_enclosure_size(main["종류"], main["극수"], main["용량"], self.branches, style="경제형")
            lines,_ = build_estimate_lines(enclosure, client, main, self.branches, self.accessories, self.var_spd.get(), style="경제형")
        except Exception as e:
            messagebox.showerror("에러", f"견적 계산 실패:\n{e}"); traceback.print_exc(); return
        self.lines=lines; self.last_enclosure={"W":W,"H":H,"D":D, **enclosure}; self.last_main=main; self.last_client=client

        title = f"[{mode.upper()}] 외함 {W}x{H}x{D} | 메인 {main['종류']} {main['극수']} {main['용량']} x {main['수량']} | 분기 {len(self.branches)}"
        self.var_est_title.set(title)
        self._render_estimate_table(lines)

        data={"timestamp": time.strftime("%Y-%m-%d %H:%M:%S"), "mode":mode, "client":client, "enclosure": self.last_enclosure,
              "main": main, "branches": self.branches, "accessories": self.accessories, "spd": self.var_spd.get(), "lines": lines,
              "status": {"order": False, "ship": False, "pay": "미입금"},
              "memo": ""}
        fname=f"{now_tag()}_{slug(client.get('업체명') or 'NONAME')}.json"
        fpath=os.path.join(EST_DIR,fname); save_json(fpath,data); self._repo_add_item(data,fpath)

    def _make_system_estimate(self): self._build_and_show_estimate("system")
    def _make_ai_estimate(self): self._build_and_show_estimate("ai")

    def _save_to_company_excel(self):
        try:
            from openpyxl import load_workbook
        except:
            messagebox.showerror("엑셀", "openpyxl이 필요합니다: pip install openpyxl")
            return
        tpl_path = os.path.join(DATA_DIR, "realsample.xlsx")
        if not os.path.exists(tpl_path):
            messagebox.showwarning("양식 없음", f"회사 양식이 없습니다.\n{tpl_path} 위치에 realsample.xlsx를 두세요.")
            return
        if not self.lines:
            messagebox.showwarning("저장 불가", "표에 표시된 견적 라인이 없습니다.")
            return
        wb = load_workbook(tpl_path)
        ws1 = wb.worksheets[0]
        try:
            ws1["A10"] = (self.last_client or {}).get("업체명","")
            ws1["A11"] = (self.last_client or {}).get("연락처","")
            ws1["A12"] = (self.last_client or {}).get("이메일","")
        except Exception:
            pass
        ws2 = wb.worksheets[1]
        row = 2
        for L in self.lines:
            ws2.cell(row=row, column=1, value=L.get("no"))
            ws2.cell(row=row, column=2, value=L.get("품명"))
            ws2.cell(row=row, column=3, value=L.get("규격"))
            ws2.cell(row=row, column=4, value=L.get("수량"))
            ws2.cell(row=row, column=5, value=L.get("단가"))
            ws2.cell(row=row, column=6, value=L.get("금액"))
            ws2.cell(row=row, column=7, value=L.get("비고",""))
            row += 1
        out = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                           filetypes=[("Excel",".xlsx")],
                                           initialfile=f"견적_{slug((self.last_client or {}).get('업체명') or '')}_{now_tag()}.xlsx")
        if not out:
            return
        try:
            wb.save(out)
            messagebox.showinfo("저장 완료", out)
        except Exception as e:
            messagebox.showerror("저장 실패", str(e))

    # ========= [탭2] AI도우미 =========
    def _build_tab_ai(self):
        tab = ttk.Frame(self.nb, padding=8)
        self.nb.add(tab, text="AI도우미")

        # 상단: 모델 선택 + 파일 첨부
        top = ttk.LabelFrame(tab, text="모델/첨부", padding=8)
        top.pack(fill="x")

        self.var_ai_model = tk.StringVar(value="chatgpt")
        ttk.Radiobutton(top, text="ChatGPT", variable=self.var_ai_model, value="chatgpt").grid(row=0, column=0, sticky="w")
        ttk.Radiobutton(top, text="Claude",  variable=self.var_ai_model, value="claude").grid(row=0, column=1, sticky="w")
        ttk.Button(top, text="파일 첨부", command=self._ai_attach_files).grid(row=0, column=2, padx=8)
        ttk.Button(top, text="RAG 재색인", command=self._rag_reindex).grid(row=0, column=3, padx=6)

        self.lst_ai_files = tk.Listbox(top, height=5, width=90)
        self.lst_ai_files.grid(row=1, column=0, columnspan=3, sticky="we", pady=(6, 0))
        if _DND_OK:
            self.lst_ai_files.drop_target_register(DND_FILES)
            self.lst_ai_files.dnd_bind('<<Drop>>', self._on_drop_ai)

        # 중간: 분석 버튼 + 메모
        mid = ttk.LabelFrame(tab, text="분석·반영", padding=8)
        mid.pack(fill="x", pady=(8, 0))
        ttk.Button(mid, text="첨부 분석 → 견적탭 자동 채움", command=self._ai_extract_to_estimate_tab).grid(row=0, column=0, padx=4)
        ttk.Label(mid, text="메모(선택)").grid(row=1, column=0, sticky="w")
        self.ent_ai_note = ttk.Entry(mid, width=90)
        self.ent_ai_note.grid(row=1, column=1, sticky="w", padx=6, pady=(4, 2))

        # 하단: 채팅창 + 입력창 + 전송
        bot = ttk.LabelFrame(tab, text="AI 채팅(업무 컨트롤)", padding=8)
        bot.pack(fill="both", expand=True, pady=(8, 0))

        self.txt_ai_chat = scrolledtext.ScrolledText(bot, wrap="word", height=18)
        self.txt_ai_chat.pack(fill="both", expand=True)

        self.txt_ai_chat.tag_configure('ai', background='#F1F5FB', foreground='#0D3B66',
                                       lmargin1=12, lmargin2=12, rmargin=80, spacing1=4, spacing3=8, justify='left')
        self.txt_ai_chat.tag_configure('user', background='#FFF5D6', foreground='#5A3E00',
                                       lmargin1=80, lmargin2=80, rmargin=12, spacing1=4, spacing3=8, justify='right')
        self.txt_ai_chat.tag_configure('sys', foreground='#666666',
                                       lmargin1=8, lmargin2=8, rmargin=8, spacing1=2, spacing3=6)
        self.txt_ai_chat.tag_config("bold", font=("맑은 고딕", 10, "bold"))

        # 읽기전용(입력 차단) + 우클릭 복사 메뉴
        def _block_keys(e): return "break"
        self.txt_ai_chat.bind("<Key>", _block_keys)
        menu = tk.Menu(self.txt_ai_chat, tearoff=0)
        menu.add_command(label="복사", command=lambda: self.txt_ai_chat.event_generate("<<Copy>>"))
        def _on_context(e):
            try: menu.tk_popup(e.x_root, e.y_root)
            finally: menu.grab_release()
        self.txt_ai_chat.bind("<Button-3>", _on_context)

        frm = ttk.Frame(bot); frm.pack(fill="x", pady=(6, 0))
        frm.grid_columnconfigure(0, weight=4)
        frm.grid_columnconfigure(1, weight=1)

        self.txt_ai_input = tk.Text(frm, height=4)
        self.txt_ai_input.grid(row=0, column=0, sticky="we")
        self.txt_ai_input.bind("<Return>", self._ai_on_enter)
        self.txt_ai_input.bind("<Shift-Return>", self._ai_on_shift_enter)

        self.btn_ai_send = ttk.Button(frm, text="보내기", command=self._ai_chat_send)
        self.btn_ai_send.grid(row=0, column=1, padx=6, sticky="e")

        # 최초 가이드 메시지
        self._ai_chat_sys("예) 'ㅇㅇ업체 견적서' → 견적서관리에서 해당 업체 검색/열람 | 'oo업체 출고건 입금했어?' → RAG에서 상태 답변")

    def _rag_reindex(self):
        try:
            if not hasattr(self, "agent") or self.agent is None:
                messagebox.showwarning("RAG", "에이전트가 초기화되지 않았습니다.")
                return
            n = self.agent.reindex()
            if n >= 0:
                messagebox.showinfo("RAG", f"인덱스 재구축 완료 (entries={n})")
            else:
                messagebox.showwarning("RAG", "재색인 실패: RAG 모듈 또는 데이터가 없습니다.")
        except Exception as e:
            messagebox.showerror("RAG", f"오류: {e}")

    # === 채팅 말풍선/전송 핸들러
    def _ai_sys_clear(self):
        try:
            self.txt_ai_chat.config(state="normal")
            start = "1.0"
            while True:
                idx = self.txt_ai_chat.search("[시스템]", start, tk.END)
                if not idx:
                    break
                line_end = f"{idx} lineend"
                self.txt_ai_chat.delete(idx, line_end+"\n")
                start = idx
            self.txt_ai_chat.config(state="disabled")
        except:
            pass

    def _ai_chat_user(self, t: str):
        self.txt_ai_chat.config(state="normal")
        self.txt_ai_chat.insert(tk.END, f"[나] {t}\n\n", ("user",))
        self.txt_ai_chat.see(tk.END)
        self.txt_ai_chat.config(state="disabled")

    def _ai_chat_bot(self, t: str):
        self.txt_ai_chat.config(state="normal")
        self.txt_ai_chat.insert(tk.END, f"[AI] {t}\n\n", ("ai",))
        self.txt_ai_chat.see(tk.END)
        self.txt_ai_chat.config(state="disabled")

    def _ai_chat_sys(self, t: str):
        self.txt_ai_chat.config(state="normal")
        self.txt_ai_chat.insert(tk.END, f"[시스템] {t}\n\n", ("sys",))
        self.txt_ai_chat.see(tk.END)
        self.txt_ai_chat.config(state="disabled")

    def _ai_on_enter(self, event):
        self._ai_chat_send()
        return "break"

    def _ai_on_shift_enter(self, event):
        self.txt_ai_input.insert(tk.INSERT, "\n")
        return "break"

    def _ai_chat_send(self):
        msg = self.txt_ai_input.get("1.0", "end-1c").strip()
        if not msg: return
        self._ai_chat_user(msg)
        self.txt_ai_input.delete("1.0", tk.END)

        if "견적서" in msg:
            name=re.sub(r"(견적서|보여줘|열어줘)","", msg).strip()
            self.nb.select(3)
            self._repo_search_and_focus(name)
            self._ai_chat_bot(f"'{name}' 관련 견적을 견적서관리에서 표시했어요.")
            return
        if "출고" in msg or "입금" in msg or "발주" in msg:
            ans=self._rag_answer_status(msg)
            self._ai_chat_bot(ans); return
        if "견적탭" in msg or "시스템 견적" in msg:
            self.nb.select(0); self._ai_chat_bot("견적 탭으로 이동했습니다."); return

        # ★ 여기부터: RAG+LLM 백그라운드 호출
        self._set_busy(True)
        self._ai_chat_sys("답변 생성 중…")
        threading.Thread(target=self._agent_worker, args=(msg,), daemon=True).start()

    def _set_busy(self, busy: bool):
        def _apply():
            try:
                self.btn_ai_send.configure(state=("disabled" if busy else "normal"))
            except:
                pass
        try:
            self.after(0, _apply)
        except Exception:
            _apply()

    def _agent_worker(self, msg: str):
        def _llm_wrapper(prompt: str) -> str:
            try:
                model = self.var_ai_model.get()
            except Exception:
                model = "chatgpt"
            return self._call_llm(model, prompt) or ""

        try:
            use_llm = bool(self.cfg.get("chatgpt",{}).get("ok") or self.cfg.get("claude",{}).get("ok"))
        except Exception:
            use_llm = False

        try:
            if hasattr(self, "agent") and self.agent is not None:
                resp = self.agent.reply(msg, llm_fn=_llm_wrapper if use_llm else None)
            else:
                resp = _llm_wrapper(msg) if use_llm else "RAG 에이전트가 초기화되지 않았습니다."
        except Exception as e:
            resp = f"처리 중 오류: {e}"

        def _deliver():
            self._ai_sys_clear()
            self._ai_chat_bot(resp or "규칙/자료 기반 응답이 없습니다.")
            self._set_busy(False)
        try:
            self.after(0, _deliver)
        except Exception:
            try:
                self.txt_ai_chat.after(0, _deliver)
            except Exception:
                _deliver()

    def _ai_attach_files(self):
        paths=filedialog.askopenfilenames(title="첨부 파일", filetypes=[("문서/이미지","*.pdf;*.png;*.jpg;*.jpeg;*.bmp;*.tif;*.tiff;*.docx;*.txt"),("모든 파일","*.*")])
        for p in paths: self.lst_ai_files.insert(tk.END,p)

    def _on_drop_ai(self, evt):
        for f in self._parse_dnd(evt.data):
            self.lst_ai_files.insert(tk.END,f)

    def _parse_dnd(self, data):
        items=[]
        for it in re.findall(r'\{[^}]+\}|[^\s]+', data):
            items.append(it.strip("{}"))
        return items

    def _ai_extract_to_estimate_tab(self):
        files=list(self.lst_ai_files.get(0,tk.END)); note=self.ent_ai_note.get().strip()
        if not files and not note:
            messagebox.showwarning("첨부 필요","파일 첨부 또는 메모 입력"); return
        tmp=None
        if note:
            tmp=os.path.join(DATA_DIR,f"_tmp_note_{now_tag()}.txt")
            with open(tmp,"w",encoding="utf-8") as f: f.write("음표")
            files=files+[tmp]
        tess=self.cfg.get("tesseract_path","")
        try:
            client, main, branches, _full=extract_info_from_files(files, tesseract_path=tess)
        except Exception as e:
            messagebox.showerror("분석 실패", f"{e}\n\n(pytesseract / pdfplumber / docx / pdf2image 중 일부가 없을 수 있습니다)")
            if tmp and os.path.exists(tmp): os.remove(tmp)
            return
        finally:
            if tmp and os.path.exists(tmp): os.remove(tmp)
        # 견적 탭 채움
        if client.get("업체명"): self.ent_client_name.delete(0,tk.END); self.ent_client_name.insert(0,client["업체명"])
        if client.get("연락처"): self.ent_client_phone.delete(0,tk.END); self.ent_client_phone.insert(0,client["연락처"])
        if client.get("이메일"): self.ent_client_email.delete(0,tk.END); self.ent_client_email.insert(0,client["이메일"])
        if main:
            self.cbo_main_kind.set(main["종류"]); self.cbo_main_poles.set(main["극수"]); self.cbo_main_amp.set(main["용량"])
            self.ent_main_qty.delete(0,tk.END); self.ent_main_qty.insert(0, main.get("수량","1"))
        self.branches.clear(); self.lst_branches.delete(0,tk.END)
        merged={}
        for b in branches:
            key=(b["종류"],b["극수"],b["용량"])
            merged[key]=merged.get(key,0)+safe_int(b.get("수량","1"),1)
        for (k,p,a),qty in merged.items():
            self.branches.append({"종류":k,"극수":p,"용량":a,"수량":str(qty)})
            self.lst_branches.insert(tk.END, f"{k} | {p} {a} | 수량 {qty}")
        self._ai_chat_sys("첨부 분석 결과를 견적 탭에 반영했습니다."); self.nb.select(0)

    # ========= [탭3] 이메일(3-Pane) =========
    def _build_tab_email(self):
        tab=ttk.Frame(self.nb, padding=8); self.nb.add(tab, text="이메일")
        # 상단 툴바
        bar=ttk.Frame(tab); bar.pack(fill="x")
        ttk.Label(bar,text="최근(일)").pack(side="left", padx=(0,4))
        self.ent_mail_days = ttk.Entry(bar, width=4); self.ent_mail_days.insert(0,"5"); self.ent_mail_days.pack(side="left")
        ttk.Button(bar,text="메일함 동기화", command=self._mail_sync).pack(side="left", padx=(4,0))
        ttk.Button(bar,text="더 불러오기(+5일)", command=self._mail_load_more).pack(side="left", padx=6)
        ttk.Button(bar,text="첨부 다운로드", command=self._mail_download).pack(side="left", padx=6)
        ttk.Button(bar,text="첨부 → AI도우미", command=self._mail_attach_to_ai).pack(side="left", padx=6)
        ttk.Button(bar,text="첨부 → AI견적", command=self._mail_attach_to_ai_est).pack(side="left", padx=6)
        ttk.Button(bar,text="메일 보내기(모의)", command=self._mail_send).pack(side="left", padx=6)

        # 3 Pane
        main=ttk.Panedwindow(tab, orient="horizontal"); main.pack(fill="both", expand=True, pady=(8,0))

        # 좌: 메일함/필터
        left=ttk.Frame(main, width=220); main.add(left, weight=1)
        ttk.Label(left,text="메일함").pack(anchor="w")
        self.lst_mailboxes=tk.Listbox(left, height=8)
        for b in ["INBOX"]:
            self.lst_mailboxes.insert(tk.END, b)
        self.lst_mailboxes.pack(fill="x")
        ttk.Label(left,text="검색").pack(anchor="w", pady=(8,2))
        self.ent_mail_search=ttk.Entry(left); self.ent_mail_search.pack(fill="x")
        ttk.Button(left,text="검색(모의)", command=self._mail_search).pack(anchor="w", pady=(6,0))

        # 중: 메일 리스트
        mid=ttk.Frame(main, width=420); main.add(mid, weight=2)
        cols=("date","from","subject","attachments")
        self.tree_mail=ttk.Treeview(mid, columns=cols, show="headings", height=20)
        for c,w in zip(cols,["날짜","보낸사람","제목","첨부"]):
            self.tree_mail.heading(c, text=w)
        self.tree_mail.column("date", width=150); self.tree_mail.column("from", width=170); self.tree_mail.column("subject", width=380)
        self.tree_mail.column("attachments", width=80, anchor="center")
        self.tree_mail.pack(fill="both", expand=True)
        self.tree_mail.bind("<<TreeviewSelect>>", self._mail_on_select)

        # 우: 본문/첨부/액션
        right=ttk.Frame(main); main.add(right, weight=3)
        self.txt_mail_body=scrolledtext.ScrolledText(right, wrap="word", height=18); self.txt_mail_body.pack(fill="both",expand=True)
        atta_bar=ttk.Frame(right); atta_bar.pack(fill="x", pady=(6,0))
        ttk.Label(atta_bar,text="첨부파일:").pack(side="left")
        self.lst_mail_attachments=tk.Listbox(atta_bar,height=4); self.lst_mail_attachments.pack(side="left", fill="x", expand=True)
        btns=ttk.Frame(right); btns.pack(fill="x", pady=(6,0))
        ttk.Button(btns,text="선택 첨부 → AI도우미", command=self._attachment_to_ai).pack(side="left")
        ttk.Button(btns,text="선택 첨부 → AI견적", command=self._mail_attach_to_ai_est).pack(side="left", padx=6)

    def _imap_since_query(self, days: int) -> str:
        import datetime
        dt = datetime.datetime.now() - datetime.timedelta(days=max(0, days))
        return dt.strftime("%d-%b-%Y")

    def _mail_load_range(self, days: int):
        try:
            imcfg   = self.cfg.get("imap", {})
            host    = imcfg.get("host", "imap.naver.com")
            port    = int(imcfg.get("port", 993))
            user    = imcfg.get("user", "").strip()
            pw      = imcfg.get("pass", "").strip()
            use_ssl = bool(imcfg.get("ssl", True))
            if not (host and user and pw):
                messagebox.showwarning("IMAP", "설정 탭에서 네이버웍스 IMAP 정보를 저장하세요.")
                return

            M = imaplib.IMAP4_SSL(host, port) if use_ssl else imaplib.IMAP4(host, port)
            M.login(user, pw)
            typ, _ = M.select("INBOX")
            if typ != "OK":
                raise RuntimeError("IMAP 메일함 선택 실패")

            since = self._imap_since_query(days)
            typ, data = M.search(None, f'(SINCE "{since}")')
            if typ != "OK":
                raise RuntimeError("IMAP 검색 실패")
            uids = list(reversed(data[0].split()))[:300]  # 최신 300개 제한

            self.tree_mail.delete(*self.tree_mail.get_children())
            self.mail_cache.clear()

            for uid in uids:
                typ, d = M.fetch(uid, "(RFC822.HEADER)")
                if typ != "OK":
                    continue
                msg = email.message_from_bytes(d[0][1])
                date_s, frm, subj, atts = _msg_summary(msg)
                self.tree_mail.insert("", "end", iid=uid.decode(),
                                      values=(date_s, frm, subj, "O" if atts else ""))
            M.logout()
        except Exception as ex:
            traceback.print_exc()
            messagebox.showerror("IMAP 오류", str(ex))

    def _mail_load_more(self):
        try:
            cur = int(self.ent_mail_days.get().strip() or "5")
        except:
            cur = 5
        cur += 5
        self.ent_mail_days.delete(0,tk.END); self.ent_mail_days.insert(0,str(cur))
        self._mail_load_range(cur)

    def _mail_sync(self):
        try:
            days = int(self.ent_mail_days.get().strip() or "5")
        except:
            days = 5
        self._mail_load_range(days)
        messagebox.showinfo("메일", f"최근 {days}일 메일함 동기화 완료")

    def _html_to_text(self, html_s: str) -> str:
        import re, html as ihtml
        s = html_s or ""
        s = re.sub(r'(?is)<(script|style).*?>.*?</\1>', ' ', s)
        s = re.sub(r'(?i)<br\s*/?>', '\n', s)
        s = re.sub(r'(?i)</p\s*>', '\n', s)
        s = re.sub(r'(?i)<(p|div|tr|h[1-6]|li|table|section|article|header|footer)[^>]*>', '\n', s)
        s = re.sub(r'(?s)<[^>]+>', ' ', s)
        s = ihtml.unescape(s).replace('\xa0', ' ')
        s = re.sub(r'[ \t]+', ' ', s)
        s = re.sub(r'\n[ \t]+', '\n', s)
        s = re.sub(r'\n{3,}', '\n\n', s)
        return s.strip()

    def _extract_best_body_text(self, msg) -> str:
        try:
            ctype = (msg.get_content_type() or "").lower()
            if not msg.is_multipart():
                raw = msg.get_payload(decode=True) or b""
                text = raw.decode(msg.get_content_charset() or "utf-8", errors="ignore")
                return text if ctype == "text/plain" else self._html_to_text(text)
            plains, htmls = [], []
            for part in msg.walk():
                cdisp = (part.get_content_disposition() or "").lower()
                if cdisp == "attachment":
                    continue
                ptype = (part.get_content_type() or "").lower()
                raw = part.get_payload(decode=True) or b""
                text = raw.decode(part.get_content_charset() or "utf-8", errors="ignore")
                if ptype == "text/plain":
                    plains.append(text)
                elif ptype == "text/html":
                    htmls.append(self._html_to_text(text))
            if plains:
                return "\n".join(plains).strip()
            if htmls:
                return "\n".join(htmls).strip()
            raw = msg.get_payload(decode=True) or b""
            return raw.decode(msg.get_content_charset() or "utf-8", errors="ignore") or "(본문이 없습니다.)"
        except Exception:
            return "(본문이 없습니다.)"

    def _mail_on_select(self, _e=None):
        sel = self.tree_mail.selection()
        if not sel:
            return
        uid = sel[0]

        imcfg   = self.cfg.get("imap", {})
        host    = imcfg.get("host", "imap.naver.com")
        port    = int(imcfg.get("port", 993))
        user    = imcfg.get("user", "").strip()
        pw      = imcfg.get("pass", "").strip()
        use_ssl = bool(imcfg.get("ssl", True))
        if not (host and user and pw):
            messagebox.showwarning("IMAP", "설정 탭에서 네이버웍스 IMAP 정보를 저장하세요.")
            return

        try:
            M = imaplib.IMAP4_SSL(host, port) if use_ssl else imaplib.IMAP4(host, port)
            M.login(user, pw)
            M.select("INBOX")

            typ, d = M.fetch(uid.encode(), "(RFC822)")
            if typ != "OK":
                raise RuntimeError("본문 로드 실패")
            msg = email.message_from_bytes(d[0][1])

            body_text = self._extract_best_body_text(msg)
            self.txt_mail_body.delete("1.0", tk.END)
            self.txt_mail_body.insert(tk.END, body_text)

            self.lst_mail_attachments.delete(0, tk.END)
            self.mail_cache[uid] = {"msg": msg}
            for part in msg.walk():
                if (part.get_content_disposition() or "").lower() == "attachment":
                    fn = _dh(part.get_filename() or "attachment.bin")
                    self.lst_mail_attachments.insert(tk.END, fn)

            M.logout()

        except Exception as ex:
            traceback.print_exc()
            messagebox.showerror("IMAP 오류", str(ex))

    def _mail_download(self):
        sel=self.tree_mail.selection()
        if not sel: return
        uid=sel[0]
        entry = self.mail_cache.get(uid)
        if not entry:
            messagebox.showwarning("첨부","먼저 메일을 선택해 본문을 로드해주세요.")
            return
        msg = entry.get("msg")
        outdir = os.path.join(DATA_DIR, "mail_attachments", uid)
        os.makedirs(outdir, exist_ok=True)
        saved=[]
        for part in msg.walk():
            if (part.get_content_disposition() or "").lower()=="attachment":
                fn=_dh(part.get_filename() or f"att_{len(saved)+1}.bin")
                path=os.path.join(outdir, fn)
                with open(path,"wb") as f:
                    f.write(part.get_payload(decode=True))
                saved.append(path)
        messagebox.showinfo("다운로드", f"{len(saved)}개 저장됨\n{outdir}")

    def _mail_send(self): messagebox.showinfo("보내기","메일 발송(모의)")
    def _mail_search(self): messagebox.showinfo("검색","검색 결과 반영(모의)")

    def _mail_attach_to_ai(self):
        sel=self.tree_mail.selection()
        if not sel: return
        uid=sel[0]
        entry = self.mail_cache.get(uid)
        if not entry:
            messagebox.showwarning("첨부","먼저 본문을 로드해주세요."); return
        msg = entry.get("msg")
        names=[]
        outdir = os.path.join(DATA_DIR, "mail_attachments", uid)
        os.makedirs(outdir, exist_ok=True)
        i=0
        for part in msg.walk():
            if (part.get_content_disposition() or "").lower()=="attachment":
                i+=1
                fn=_dh(part.get_filename() or f"att_{i}.bin")
                path=os.path.join(outdir, fn)
                with open(path,"wb") as f: f.write(part.get_payload(decode=True))
                names.append(path)
        self.nb.select(1)
        for p in names: self.lst_ai_files.insert(tk.END, p)
        messagebox.showinfo("전송", f"AI도우미로 {len(names)}개 보냈습니다.")

    def _mail_attach_to_ai_est(self):
        # 선택 첨부 저장 + AI도우미 첨부 리스트에 추가
        self._attachment_to_ai()
        # → 첨부 분석을 통해 견적 탭 자동 채움 시도
        try:
            self.nb.select(1)  # AI도우미 탭
            self._ai_extract_to_estimate_tab()   # 첨부 분석 → 견적탭 채움
            self.nb.select(0)  # 견적 탭
            messagebox.showinfo("완료","첨부를 분석하여 견적 탭에 반영했습니다.")
        except Exception as e:
            messagebox.showerror("실패", f"AI 견적 자동 반영 실패: {e}")

    def _attachment_to_ai(self):
        sels=self.lst_mail_attachments.curselection()
        if not sels: return
        uid = self.tree_mail.selection()[0]
        entry = self.mail_cache.get(uid)
        if not entry:
            messagebox.showwarning("첨부","먼저 본문을 로드해주세요."); return
        msg = entry.get("msg")
        outdir = os.path.join(DATA_DIR, "mail_attachments", uid)
        os.makedirs(outdir, exist_ok=True)
        names=[]
        i=0
        for part in msg.walk():
            if (part.get_content_disposition() or "").lower()=="attachment":
                i+=1
                fn=_dh(part.get_filename() or f"att_{i}.bin")
                if (i-1) in sels:
                    path=os.path.join(outdir, fn)
                    with open(path,"wb") as f: f.write(part.get_payload(decode=True))
                    names.append(path)
        self.nb.select(1)
        for p in names: self.lst_ai_files.insert(tk.END, p)
        messagebox.showinfo("전송","AI도우미로 보냈습니다.")

    # ========= [탭4] 견적서 관리 =========
    def _build_tab_repo(self):
        tab=ttk.Frame(self.nb, padding=8); self.nb.add(tab, text="견적서관리")
        top=ttk.Frame(tab); top.pack(fill="x")
        ttk.Button(top,text="새로고침", command=self._repo_reload).pack(side="left")
        ttk.Button(top,text="Excel 저장(.xlsx)", command=self._repo_save_excel).pack(side="left",padx=6)
        ttk.Button(top,text="CSV 저장(.csv)", command=self._repo_save_csv).pack(side="left")
        ttk.Label(top,text="상태:").pack(side="left", padx=(18,4))
        self.var_repo_order = tk.BooleanVar(value=False)
        self.var_repo_ship  = tk.BooleanVar(value=False)
        self.var_repo_pay   = tk.StringVar(value="미입금")
        ttk.Checkbutton(top,text="발주", variable=self.var_repo_order, command=self._repo_update_status).pack(side="left")
        ttk.Checkbutton(top,text="출고", variable=self.var_repo_ship,  command=self._repo_update_status).pack(side="left",padx=(6,0))
        ttk.Combobox(top, values=["미입금","계약금","중도금","잔금","완납"], textvariable=self.var_repo_pay, width=8, state="readonly").pack(side="left", padx=(6,0))
        ttk.Label(top,text="검색 업체").pack(side="left", padx=(18,4))
        self.ent_repo_search=ttk.Entry(top, width=20); self.ent_repo_search.pack(side="left")
        ttk.Button(top,text="찾기", command=lambda:self._repo_search_and_focus(self.ent_repo_search.get().strip())).pack(side="left", padx=4)

        cols=("time","client","order","ship","pay","memo","path")
        self.tree_repo=ttk.Treeview(tab, columns=cols, show="headings", height=22)
        headers=["생성시각","업체명","발주","출고","입금","비고","(경로숨김)"]
        widths =[140,220,60,60,100,260,10]
        for c,h,w in zip(cols,headers,widths):
            self.tree_repo.heading(c, text=h)
            self.tree_repo.column(c, width=w)
        self.tree_repo.column("path", width=1, stretch=False)
        self.tree_repo.pack(fill="both", expand=True, pady=(8,0))
        self.tree_repo.bind("<<TreeviewSelect>>", self._repo_on_select)
        self._repo_reload()

    def _repo_reload(self):
        self.tree_repo.delete(*self.tree_repo.get_children())
        for fn in sorted(os.listdir(EST_DIR)):
            if not fn.lower().endswith(".json"): continue
            p=os.path.join(EST_DIR,fn); data=load_json(p,{})
            client=(data.get("client") or {}).get("업체명",""); ts=data.get("timestamp","")
            st=data.get("status",{"order":False,"ship":False,"pay":"미입금"})
            memo=data.get("memo","")
            self.tree_repo.insert("", "end", values=(
                ts, client,
                "O" if st.get("order") else "",
                "O" if st.get("ship") else "",
                st.get("pay","미입금"),
                memo, p
            ))

    def _repo_on_select(self,_=None):
        item=self._repo_selected_item()
        if not item: return
        data=load_json(item["path"],{})
        st=data.get("status", {"order":False,"ship":False,"pay":"미입금"})
        self.var_repo_order.set(bool(st.get("order",False))); self.var_repo_ship.set(bool(st.get("ship",False))); self.var_repo_pay.set(st.get("pay","미입금"))

    def _repo_selected_item(self):
        sel=self.tree_repo.selection()
        if not sel: return None
        v=self.tree_repo.item(sel[0],"values")
        return {"time":v[0], "client":v[1], "path":v[6]}

    def _repo_update_status(self):
        it=self._repo_selected_item()
        if not it: return
        d=load_json(it["path"],{})
        d["status"]={"order": self.var_repo_order.get(), "ship":self.var_repo_ship.get(), "pay":self.var_repo_pay.get()}
        save_json(it["path"], d); messagebox.showinfo("저장","상태 저장")
        self._repo_reload()

    def _repo_add_item(self, data, path):
        st=data.get("status",{"order":False,"ship":False,"pay":"미입금"})
        self.tree_repo.insert("", "end", values=(
            data["timestamp"], data["client"].get("업체명",""),
            "O" if st.get("order") else "",
            "O" if st.get("ship") else "",
            st.get("pay","미입금"),
            data.get("memo",""),
            path
        ))

    def _repo_save_excel(self):
        it=self._repo_selected_item()
        if not it: messagebox.showwarning("선택 없음","저장할 견적 선택"); return
        from openpyxl import Workbook
        d=load_json(it["path"],{})
        wb=Workbook(); ws=wb.active; ws.title="견적"
        ws.append(["No","품명","규격","단위","수량","단가","금액"])
        for L in d.get("lines",[]): ws.append([L["no"],L["품명"],L["규격"],L["단위"],L["수량"],L["단가"],L["금액"]])
        out=filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel",".xlsx")],
                                         initialfile=f"repo_{now_tag()}.xlsx")
        if not out: return
        wb.save(out); messagebox.showinfo("저장", out)

    def _repo_save_csv(self):
        it=self._repo_selected_item()
        if not it: 
            messagebox.showwarning("선택 없음","저장할 견적 선택"); 
            return
        d=load_json(it["path"],{})
        out=filedialog.asksaveasfilename(defaultextension=".csv",
                                         filetypes=[("CSV",".csv")],
                                         initialfile=f"repo_{now_tag()}.csv")
        if not out: return
        import csv
        with open(out,"w",newline="",encoding="utf-8-sig") as f:
            w=csv.writer(f); w.writerow(["No","품명","규격","단위","수량","단가","금액","비고"])
            for L in d.get("lines",[]): 
                w.writerow([L.get("no",""),L.get("품명",""),L.get("규격",""),
                            L.get("단위",""),L.get("수량",""),L.get("단가",""),
                            L.get("금액",""),L.get("비고","")])
        messagebox.showinfo("저장", out)

    # ========= [탭5] 단가관리 =========
    def _build_tab_prices(self):
        tab=ttk.Frame(self.nb, padding=8); self.nb.add(tab, text="단가관리")
        bar=ttk.Frame(tab); bar.pack(fill="x")
        ttk.Button(bar,text="외함 단가 첨부", command=lambda:self._price_attach("enclosure")).pack(side="left")
        ttk.Button(bar,text="차단기 단가 첨부", command=lambda:self._price_attach("breaker")).pack(side="left", padx=6)
        ttk.Button(bar,text="부속자재 단가 첨부", command=lambda:self._price_attach("accessory")).pack(side="left")
        ttk.Label(bar,text="  (여기에 저장: data/prices)").pack(side="left", padx=8)

        self.txt_prices=scrolledtext.ScrolledText(tab, wrap="none", height=24); 
        self.txt_prices.pack(fill="both",expand=True,pady=(8,0))
        self._load_prices_to_editor()

        info=ttk.Label(tab, text="※ 메커니즘: 이 탭은 파일을 data/prices 폴더에 보관합니다.\n"
                                 "   - 파일명 접두어로 종류를 구분: enclosure_*, breaker_*, accessory_*\n"
                                 "   - 이후 매칭 모듈이 최신 파일을 읽어 단가 조회(RAG/룰) 기반으로 사용합니다.",
                       justify="left")
        info.pack(fill="x", pady=(6,0))

    def _price_attach(self, kind):
        path=filedialog.askopenfilename(title="단가 파일", 
                                        filetypes=[("Excel/CSV","*.xlsx;*.csv"),("모든 파일","*.*")])
        if not path: return
        dest=os.path.join(PRICES_DIR, f"{kind}_{os.path.basename(path)}")
        try:
            import shutil; shutil.copy2(path, dest)
        except Exception as e: 
            messagebox.showerror("오류", f"복사 실패: {e}"); 
            return
        messagebox.showinfo("저장", f"{dest} 저장")
        self._load_prices_to_editor()

    def _load_prices_to_editor(self):
        self.txt_prices.delete("1.0",tk.END)
        self.txt_prices.insert(tk.END, "# 단가 파일 목록\n")
        files=sorted(os.listdir(PRICES_DIR)) if os.path.exists(PRICES_DIR) else []
        if not files:
            self.txt_prices.insert(tk.END, "- (없음)\n")
            return
        for fn in files:
            self.txt_prices.insert(tk.END, f"- {fn}\n")

    # ========= [탭6] 거래처관리 (모달로 추가 + 분석 요약) =========
    def _build_tab_clients(self):
        tab=ttk.Frame(self.nb, padding=8); self.nb.add(tab, text="거래처관리")
        top=ttk.Frame(tab); top.pack(fill="x")
        ttk.Button(top, text="거래처 추가", command=self._client_add_modal).pack(side="left")
        ttk.Label(top, text="  검색:").pack(side="left", padx=(10,4))
        self.ent_client_search=ttk.Entry(top, width=24); self.ent_client_search.pack(side="left")
        ttk.Button(top, text="찾기", command=self._clients_reload).pack(side="left", padx=4)

        cols=("name","owner","phone","email","biz","type","stats")
        self.tree_clients=ttk.Treeview(tab, columns=cols, show="headings", height=20)
        headers=["업체명","대표자","연락처","이메일","사업자번호","구분(매출/매입)","분석 요약"]
        widths =[180,80,120,200,120,120,460]
        for c,h,w in zip(cols,headers,widths):
            self.tree_clients.heading(c, text=h)
            self.tree_clients.column(c, width=w)
        self.tree_clients.pack(fill="both", expand=True, pady=(8,0))
        self._clients_reload()

    def _client_add_modal(self):
        win=tk.Toplevel(self); win.title("거래처 추가"); win.grab_set()
        frm=ttk.Frame(win, padding=12); frm.pack(fill="both", expand=True)
        def row(r,lbl,w=26,show=None):
            ttk.Label(frm,text=lbl).grid(row=r,column=0,sticky="e",pady=2)
            e=ttk.Entry(frm,width=w, show=show); e.grid(row=r,column=1,sticky="w",pady=2)
            return e
        e_name=row(0,"업체명*"); e_owner=row(1,"대표자"); e_phone=row(2,"연락처")
        e_email=row(3,"이메일", w=36); e_addr=row(4,"주소", w=36)
        e_fax=row(5,"팩스"); e_biz=row(6,"사업자등록번호")
        ttk.Label(frm,text="구분").grid(row=7,column=0,sticky="e",pady=2)
        cbo_type=ttk.Combobox(frm, values=["매출처","매입처","겸용"], state="readonly", width=10); 
        cbo_type.set("매출처"); cbo_type.grid(row=7,column=1,sticky="w",pady=2)
        ttk.Label(frm,text="비고").grid(row=8,column=0,sticky="e",pady=2)
        e_note=ttk.Entry(frm, width=36); e_note.grid(row=8,column=1,sticky="w",pady=2)

        btns=ttk.Frame(frm); btns.grid(row=9,column=0,columnspan=2, pady=(10,0))
        def _ok():
            name=e_name.get().strip()
            if not name:
                messagebox.showwarning("필수","업체명은 필수입니다.")
                return
            db=load_json(CLIENTS_JSON, default={"clients":[]})
            db["clients"].append({
                "name":name, "owner":e_owner.get().strip(), "phone":e_phone.get().strip(),
                "email":e_email.get().strip(), "addr":e_addr.get().strip(), "fax":e_fax.get().strip(),
                "biz":e_biz.get().strip(), "type":cbo_type.get(), "note":e_note.get().strip()
            })
            save_json(CLIENTS_JSON, db)
            self._clients_reload()
            win.destroy()
        ttk.Button(btns,text="확인", command=_ok).pack(side="left", padx=6)
        ttk.Button(btns,text="취소", command=win.destroy).pack(side="left")

    def _clients_reload(self):
        db=load_json(CLIENTS_JSON, default={"clients":[]})
        q=(self.ent_client_search.get().strip() if hasattr(self,"ent_client_search") else "")
        self.tree_clients.delete(*self.tree_clients.get_children())
        for c in db["clients"]:
            if q and q not in c.get("name",""):
                continue
            stat=self._calc_client_stats(c.get("name"))
            summary=(f"견적 {stat['estimate_count']}건, 발주 {stat['order_count']}회, "
                     f"주문율 {stat['order_rate']}%, 최근6M {stat['sales_6m']:,}원, 최근1Y {stat['sales_1y']:,}원, "
                     f"입금상태 {stat['last_pay'] or '미입금'}")
            self.tree_clients.insert("", "end", values=(
                c.get("name",""), c.get("owner",""), c.get("phone",""),
                c.get("email",""), c.get("biz",""), c.get("type",""), summary
            ))

    def _calc_client_stats(self, name=None):
        # 간단 통계 (파일 타임스탬프 기반 최근 매출 집계)
        import datetime
        est_files=[os.path.join(EST_DIR,fn) for fn in os.listdir(EST_DIR) if fn.endswith(".json")]
        ests=[load_json(p,{}) for p in est_files]
        if name: ests=[e for e in ests if (e.get("client",{}).get("업체명","")==name)]
        if not ests and name: 
            return {"name":name,"estimate_count":0,"order_count":0,"order_rate":0,"sales_6m":0,"sales_1y":0,"last_pay":None}
        order_count=0; last_pay=None
        sales_6m=sales_1y=0; now=datetime.datetime.now()
        for e in ests:
            st=e.get("status",{})
            if st.get("order"): order_count+=1
            if st.get("pay"): last_pay=st.get("pay")
            # 금액 합계
            subtotal=0
            for L in e.get("lines",[]):
                try: subtotal+=int(float(L.get("금액",0)))
                except: pass
            # 기간 분배 (파일에 timestamp가 문자열로 저장되어 있음)
            ts=e.get("timestamp","")
            try:
                dt=datetime.datetime.strptime(ts,"%Y-%m-%d %H:%M:%S")
            except:
                dt=now
            if (now-dt).days<=183: sales_6m+=subtotal
            if (now-dt).days<=365: sales_1y+=subtotal
        est_count=len(ests)
        order_rate= round((order_count/max(1,est_count))*100)
        return {"name":name or "(전체)","estimate_count":est_count,"order_count":order_count,
                "order_rate":order_rate,"sales_6m":sales_6m,"sales_1y":sales_1y,"last_pay":last_pay}

    # ========= [탭7] 발주·출고 (골격) =========
    def _build_tab_shipping(self):
        tab=ttk.Frame(self.nb, padding=8); self.nb.add(tab, text="발주·출고")
        ttk.Label(tab,text="발주/출고 진행 현황 (개발 중)").pack(anchor="w")

    # ========= [탭8] 보고서 =========
    def _build_tab_reports(self):
        tab=ttk.Frame(self.nb, padding=8); self.nb.add(tab, text="보고서")
        frm=ttk.LabelFrame(tab, text="보고서 생성", padding=8); frm.pack(fill="x")
        ttk.Label(frm,text="기간(예: 2025-08-01~2025-08-31)").grid(row=0,column=0)
        self.ent_report_range=ttk.Entry(frm, width=28); self.ent_report_range.grid(row=0,column=1,padx=6)
        ttk.Button(frm,text="일일", command=lambda:self._report_make("daily")).grid(row=0,column=2,padx=4)
        ttk.Button(frm,text="주간", command=lambda:self._report_make("weekly")).grid(row=0,column=3,padx=4)
        ttk.Button(frm,text="월간", command=lambda:self._report_make("monthly")).grid(row=0,column=4,padx=4)
        self.txt_report=scrolledtext.ScrolledText(tab, wrap="word", height=22); self.txt_report.pack(fill="both",expand=True,pady=(8,0))

    def _report_make(self, kind):
        self.txt_report.delete("1.0",tk.END)
        self.txt_report.insert(tk.END, f"[{kind}] 보고서(모의)\n- 기간: {(self.ent_report_range.get().strip() or '지정 없음')}\n")
        cnt=len([fn for fn in os.listdir(EST_DIR) if fn.endswith('.json')])
        self.txt_report.insert(tk.END, f"- 누적 견적: {cnt} 건\n"); self.txt_report.see(tk.END)

    # ========= [탭9] 설정 =========
    def _build_tab_settings(self):
        tab=ttk.Frame(self.nb, padding=8); self.nb.add(tab, text="설정")

        # SMTP(좌) + 네이버웍스(우)
        top=ttk.Frame(tab); top.pack(fill="x")
        g1=ttk.LabelFrame(top, text="SMTP 메일", padding=8); g1.pack(side="left", fill="x", expand=True, padx=(0,6))
        self.ent_smtp_host=ttk.Entry(g1,width=22); self.ent_smtp_port=ttk.Entry(g1,width=6)
        self.ent_smtp_user=ttk.Entry(g1,width=22); self.ent_smtp_pass=ttk.Entry(g1,width=22, show="*")
        self.ent_smtp_sender=ttk.Entry(g1,width=28)
        ttk.Label(g1,text="Host").grid(row=0,column=0); self.ent_smtp_host.grid(row=0,column=1,padx=4)
        ttk.Label(g1,text="Port").grid(row=0,column=2); self.ent_smtp_port.grid(row=0,column=3,padx=4)
        ttk.Label(g1,text="User").grid(row=1,column=0); self.ent_smtp_user.grid(row=1,column=1,padx=4)
        ttk.Label(g1,text="Pass").grid(row=1,column=2); self.ent_smtp_pass.grid(row=1,column=3,padx=4)
        ttk.Label(g1,text="Sender").grid(row=2,column=0); self.ent_smtp_sender.grid(row=2,column=1,padx=4, columnspan=3, sticky="we")

        gIMAP = ttk.LabelFrame(tab, text="네이버웍스 메일(IMAP 인증)", padding=8); gIMAP.pack(fill="x", pady=(8, 0))
        self.ent_imap_host = ttk.Entry(gIMAP, width=22)
        self.ent_imap_port = ttk.Entry(gIMAP, width=6)
        self.ent_imap_user = ttk.Entry(gIMAP, width=26)
        self.ent_imap_pass = ttk.Entry(gIMAP, width=26, show="*")
        self.var_imap_ssl  = tk.BooleanVar(value=True)
        ttk.Label(gIMAP, text="Host").grid(row=0, column=0, sticky="e"); self.ent_imap_host.grid(row=0, column=1, padx=4, sticky="w")
        ttk.Label(gIMAP, text="Port").grid(row=0, column=2, sticky="e"); self.ent_imap_port.grid(row=0, column=3, padx=4, sticky="w")
        ttk.Label(gIMAP, text="User").grid(row=1, column=0, sticky="e"); self.ent_imap_user.grid(row=1, column=1, padx=4, sticky="w")
        ttk.Label(gIMAP, text="App Password").grid(row=1, column=2, sticky="e"); self.ent_imap_pass.grid(row=1, column=3, padx=4, sticky="w")
        ttk.Checkbutton(gIMAP, text="SSL(권장)", variable=self.var_imap_ssl).grid(row=0, column=4, padx=8)
        ttk.Button(gIMAP, text="IMAP 연결테스트", command=self._test_imap).grid(row=1, column=4, padx=8)

        gNW=ttk.LabelFrame(top, text="네이버웍스 (ID + API Key)", padding=8); gNW.pack(fill="x", pady=(8,0))
        self.ent_nw_id=ttk.Entry(gNW, width=28); self.ent_nw_key=ttk.Entry(gNW, width=28, show="*")
        ttk.Label(gNW,text="ID").grid(row=0,column=0); self.ent_nw_id.grid(row=0,column=1,padx=4)
        ttk.Label(gNW,text="API Key").grid(row=1,column=0); self.ent_nw_key.grid(row=1,column=1,padx=4)

        # 외부 서비스(Only ChatGPT/Claude)
        g2=ttk.LabelFrame(tab, text="외부 서비스 (AI)", padding=8); g2.pack(fill="x", pady=(8,0))
        ttk.Label(g2,text="ChatGPT URL").grid(row=0,column=0,sticky="e"); 
        self.ent_gpt_url=ttk.Entry(g2,width=48); self.ent_gpt_url.grid(row=0,column=1,padx=4,sticky="w")
        ttk.Label(g2,text="API Key").grid(row=0,column=2,sticky="e"); 
        self.ent_gpt_key=ttk.Entry(g2,width=36, show="*"); self.ent_gpt_key.grid(row=0,column=3,padx=4,sticky="w")
        ttk.Button(g2,text="연결테스트", command=lambda:self._test_ai("chatgpt")).grid(row=0,column=4,padx=6)

        ttk.Label(g2,text="Claude URL").grid(row=1,column=0,sticky="e"); 
        self.ent_claude_url=ttk.Entry(g2,width=48); self.ent_claude_url.grid(row=1,column=1,padx=4,sticky="w")
        ttk.Label(g2,text="API Key").grid(row=1,column=2,sticky="e"); 
        self.ent_claude_key=ttk.Entry(g2,width=36, show="*"); self.ent_claude_key.grid(row=1,column=3,padx=4,sticky="w")
        ttk.Button(g2,text="연결테스트", command=lambda:self._test_ai("claude")).grid(row=1,column=4,padx=6)

        g3=ttk.LabelFrame(tab, text="로컬 도구", padding=8); g3.pack(fill="x", pady=(8,0))
        self.ent_tess=ttk.Entry(g3, width=60)
        ttk.Label(g3,text="Tesseract 경로").grid(row=0,column=0); self.ent_tess.grid(row=0,column=1, padx=4, sticky="we")

        ttk.Button(tab,text="설정 불러오기", command=self._cfg_load).pack(side="left", padx=(0,6), pady=(8,0))
        ttk.Button(tab,text="설정 저장", command=self._cfg_save).pack(side="left", pady=(8,0))
        ttk.Button(tab,text="설정 테스트(모의)", command=lambda: messagebox.showinfo("테스트","일부 모듈 테스트(모의) OK")).pack(side="left", padx=6, pady=(8,0))

        self._cfg_load()

    # ========== 공통 유틸 ==========

    def _http_session_with_retry(self):
        s = requests.Session()
        retry = Retry(total=3, backoff_factor=0.8,
                      status_forcelist=[429, 500, 502, 503, 504],
                      allowed_methods=["POST"])
        adapter = HTTPAdapter(max_retries=retry, pool_connections=4, pool_maxsize=4)
        s.mount("https://", adapter)
        s.mount("http://", adapter)
        return s

    def _call_llm(self, which: str, msg: str):
        try:
            sess = self._http_session_with_retry()
            if which=="chatgpt" and self.cfg.get("chatgpt",{}).get("ok"):
                url=self.cfg["chatgpt"]["url"]; key=self.cfg["chatgpt"]["api_key"]; model=self.cfg["chatgpt"].get("model","gpt-4o")
                headers={"Authorization": f"Bearer {key}","Content-Type":"application/json"}
                body={"model": model, "messages":[{"role":"user","content":msg}],
                      "temperature":0.2, "max_tokens":400}
                r = sess.post(url, headers=headers, json=body, timeout=(10, 60))
                if r.ok:
                    data=r.json()
                    content = data.get("choices",[{}])[0].get("message",{}).get("content")
                    return content or "응답 없음"
                try:    return f"ChatGPT 응답 실패: {r.status_code} {r.json()}"
                except: return f"ChatGPT 응답 실패: {r.status_code} {r.text[:300]}"

            if which=="claude" and self.cfg.get("claude",{}).get("ok"):
                url=self.cfg["claude"]["url"]; key=self.cfg["claude"]["api_key"]; model=self.cfg["claude"].get("model","claude-3-7-sonnet-20250219")
                headers={"x-api-key":key,"anthropic-version":"2023-06-01","content-type":"application/json"}
                body={"model":model,"max_tokens":400,"messages":[{"role":"user","content":msg}]}
                r = sess.post(url, headers=headers, json=body, timeout=(10, 60))
                if r.ok:
                    data=r.json()
                    content = "".join([b.get("text","") for b in data.get("content",[]) if b.get("type")=="text"])
                    return content or "응답 없음"
                try:    return f"Claude 응답 실패: {r.status_code} {r.json()}"
                except: return f"Claude 응답 실패: {r.status_code} {r.text[:400]}"
        except requests.exceptions.ReadTimeout:
            return "네트워크 지연으로 응답 시간이 초과됐습니다. 잠시 후 다시 시도해주세요."
        except Exception as e:
            return f"연결 오류: {e}"
        return None

    def _test_ai(self, which: str):
        if which=="chatgpt":
            url=self.ent_gpt_url.get().strip(); key=self.ent_gpt_key.get().strip()
            if not url or not key:
                messagebox.showwarning("필수","URL과 API Key 입력"); return
            try:
                headers={"Authorization": f"Bearer {key}","Content-Type":"application/json"}
                body={"model": self.cfg["chatgpt"].get("model","gpt-4o"),
                      "messages":[{"role":"user","content":"ping"}], "max_tokens":10}
                r=requests.post(url, headers=headers, json=body, timeout=12)
                self.cfg["chatgpt"]["ok"]=bool(r.ok); save_json(CFG_PATH,self.cfg)
                if r.ok: messagebox.showinfo("성공","ChatGPT 연결 성공")
                else:
                    try: err=r.json()
                    except: err={"text": r.text[:300]}
                    messagebox.showerror("실패", f"HTTP {r.status_code}\n{err}")
            except Exception as e:
                self.cfg["chatgpt"]["ok"]=False; save_json(CFG_PATH,self.cfg)
                messagebox.showerror("오류", f"{e}")
        else:
            url=self.ent_claude_url.get().strip() or "https://api.anthropic.com/v1/messages"
            key=self.ent_claude_key.get().strip()
            if not url or not key:
                messagebox.showwarning("필수","URL과 API Key 입력"); return
            try:
                headers={"x-api-key":key,"anthropic-version":"2023-06-01","content-type":"application/json"}
                body={"model": self.cfg["claude"].get("model","claude-3-7-sonnet-20250219"),
                      "max_tokens":32,"messages":[{"role":"user","content":"ping"}]}
                r=requests.post(url, headers=headers, json=body, timeout=20)
                self.cfg["claude"]["ok"]=bool(r.ok); save_json(CFG_PATH,self.cfg)
                if r.ok: messagebox.showinfo("성공","Claude 연결 성공")
                else:
                    try: err=r.json()
                    except: err={"text": r.text[:500]}
                    messagebox.showerror("실패", f"HTTP {r.status_code}\n{err}")
            except Exception as e:
                self.cfg["claude"]["ok"]=False; save_json(CFG_PATH,self.cfg)
                messagebox.showerror("오류", f"{e}")

    def _test_imap(self):
        host = self.ent_imap_host.get().strip() or "imap.naver.com"
        port = int(self.ent_imap_port.get().strip() or "993")
        user = self.ent_imap_user.get().strip()
        pw   = self.ent_imap_pass.get().strip()
        use_ssl = bool(self.var_imap_ssl.get())
        if not (host and user and pw):
            messagebox.showwarning("IMAP", "Host / User / App Password를 입력하세요.")
            return
        try:
            M = imaplib.IMAP4_SSL(host, port) if use_ssl else imaplib.IMAP4(host, port)
            M.login(user, pw); M.select("INBOX")
            typ, _ = M.search(None, "ALL"); M.logout()
            ok = (typ=="OK")
            self.cfg.setdefault("imap", {})["ok"] = ok
            save_json(CFG_PATH, self.cfg)
            messagebox.showinfo("성공" if ok else "실패",
                                "IMAP 연결 성공 (INBOX 접근 확인)" if ok else "IMAP 검색 실패 (권한/설정 확인)")
        except Exception as e:
            self.cfg.setdefault("imap", {})["ok"] = False
            save_json(CFG_PATH, self.cfg)
            messagebox.showerror("IMAP 오류", str(e))

    def _cfg_load(self):
        cfg=self.cfg
        s=cfg.get("smtp",{})
        self.ent_smtp_host.delete(0,tk.END); self.ent_smtp_host.insert(0,s.get("host",""))
        self.ent_smtp_port.delete(0,tk.END); self.ent_smtp_port.insert(0,s.get("port",587))
        self.ent_smtp_user.delete(0,tk.END); self.ent_smtp_user.insert(0,s.get("user",""))
        self.ent_smtp_pass.delete(0,tk.END); self.ent_smtp_pass.insert(0,s.get("pass",""))
        self.ent_smtp_sender.delete(0,tk.END); self.ent_smtp_sender.insert(0,s.get("sender",""))

        im = cfg.get("imap", {})
        self.ent_imap_host.delete(0, tk.END); self.ent_imap_host.insert(0, im.get("host", "imap.naver.com"))
        self.ent_imap_port.delete(0, tk.END); self.ent_imap_port.insert(0, im.get("port", 993))
        self.ent_imap_user.delete(0, tk.END); self.ent_imap_user.insert(0, im.get("user", ""))
        self.ent_imap_pass.delete(0, tk.END); self.ent_imap_pass.insert(0, im.get("pass", ""))
        self.var_imap_ssl.set(bool(im.get("ssl", True)))

        nw=cfg.get("naver_works",{})
        self.ent_nw_id.delete(0,tk.END); self.ent_nw_id.insert(0,nw.get("id",""))
        self.ent_nw_key.delete(0,tk.END); self.ent_nw_key.insert(0,nw.get("api_key",""))

        # AI
        self.ent_gpt_url.delete(0,tk.END); self.ent_gpt_url.insert(0, cfg.get("chatgpt",{}).get("url",""))
        self.ent_gpt_key.delete(0,tk.END); self.ent_gpt_key.insert(0, cfg.get("chatgpt",{}).get("api_key",""))
        self.ent_claude_url.delete(0,tk.END); self.ent_claude_url.insert(0, cfg.get("claude",{}).get("url",""))
        self.ent_claude_key.delete(0,tk.END); self.ent_claude_key.insert(0, cfg.get("claude",{}).get("api_key",""))

        self.ent_tess.delete(0,tk.END); self.ent_tess.insert(0, cfg.get("tesseract_path",""))

    def _cfg_save(self):
        try:
            self.cfg["smtp"]={"host": self.ent_smtp_host.get().strip(),
                              "port": int(self.ent_smtp_port.get().strip() or "587"),
                              "user": self.ent_smtp_user.get().strip(),
                              "pass": self.ent_smtp_pass.get().strip(),
                              "sender": self.ent_smtp_sender.get().strip()}
            self.cfg["imap"] = {
                "host": self.ent_imap_host.get().strip() or "imap.naver.com",
                "port": int(self.ent_imap_port.get().strip() or "993"),
                "user": self.ent_imap_user.get().strip(),
                "pass": self.ent_imap_pass.get().strip(),
                "ssl":  bool(self.var_imap_ssl.get()),
                "ok":   bool(self.cfg.get("imap", {}).get("ok", False))
            }
            self.cfg["naver_works"]={"id": self.ent_nw_id.get().strip(),
                                     "api_key": self.ent_nw_key.get().strip()}
            self.cfg.setdefault("chatgpt",{})
            self.cfg["chatgpt"]["url"]=self.ent_gpt_url.get().strip()
            self.cfg["chatgpt"]["api_key"]=self.ent_gpt_key.get().strip()
            self.cfg.setdefault("claude",{})
            self.cfg["claude"]["url"]=self.ent_claude_url.get().strip()
            self.cfg["claude"]["api_key"]=self.ent_claude_key.get().strip()
            self.cfg["tesseract_path"]=self.ent_tess.get().strip()
            save_json(CFG_PATH, self.cfg); messagebox.showinfo("저장","설정 저장 완료")
        except Exception as e:
            messagebox.showerror("오류", f"설정 저장 실패: {e}")

    # ========= 고객정보 채우기(선택된 메일 기반) =========
    def _fill_client_from_selected_email(self):
        # 이메일 탭에서 선택된 메일의 본문/헤더에서 업체명/연락처/이메일 추정
        try:
            sel = getattr(self, "tree_mail").selection()
        except Exception:
            sel = None
        if not sel:
            messagebox.showwarning("안내","이메일 탭에서 먼저 메일을 선택하세요.")
            return
        uid = sel[0]
        entry = self.mail_cache.get(uid)
        if not entry:
            messagebox.showwarning("안내","선택된 메일 본문을 먼저 열어주세요.")
            return
        msg = entry.get("msg")
        body = self._extract_best_body_text(msg)
        # 헤더 + 본문 통합 텍스트에서 추출
        header_txt = f"From: {_dh(msg.get('From',''))}\nSubject: {_dh(msg.get('Subject',''))}\n"
        full = header_txt + "\n" + (body or "")
        client={"업체명":_extract_company(full),"연락처":_extract_phone(full),"이메일":_extract_email(full)}
        if client.get("업체명"): self.ent_client_name.delete(0,tk.END); self.ent_client_name.insert(0,client["업체명"])
        if client.get("연락처"): self.ent_client_phone.delete(0,tk.END); self.ent_client_phone.insert(0,client["연락처"])
        if client.get("이메일"): self.ent_client_email.delete(0,tk.END); self.ent_client_email.insert(0,client["이메일"])
        messagebox.showinfo("완료","메일에서 고객 정보를 불러왔습니다.")

# -------- 실행 --------
def main():
    app=App(); app.mainloop()
if __name__=="__main__":
    main()